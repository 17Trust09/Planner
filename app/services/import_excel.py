from __future__ import annotations

from pathlib import Path
from typing import Dict

from openpyxl import load_workbook

from app.models.definitions import GLOBAL_TOPICS, ROOM_TOPICS
from app.models.project import Project, TopicState, create_empty_project


def _import_sheet(ws, topic_by_title: Dict[str, str], states: Dict[str, TopicState]) -> None:
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        topic_title = row[1]
        if not topic_title or topic_title not in topic_by_title:
            continue
        key = topic_by_title[topic_title]
        selections_text = row[2] or ""
        notes = row[3] or ""
        assignee = row[4] if len(row) > 4 and row[4] else ""
        selections = [s.strip() for s in str(selections_text).split(",") if s.strip() and s.strip() != "—"]
        states[key] = TopicState(selections=selections[:3], notes=str(notes).strip(), assignee=str(assignee).strip())


def import_project_from_excel(path: Path, project_name: str = "Importiertes Projekt") -> Project:
    wb = load_workbook(path)
    project = create_empty_project(project_name)

    global_by_title = {t.title: t.key for t in GLOBAL_TOPICS}
    room_by_title = {t.title: t.key for t in ROOM_TOPICS}

    if "Global_Planung" in wb.sheetnames:
        _import_sheet(wb["Global_Planung"], global_by_title, project.global_topics)

    for room_name, room in project.rooms.items():
        if room_name in wb.sheetnames:
            _import_sheet(wb[room_name], room_by_title, room.topics)
    return project
