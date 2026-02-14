from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.models.definitions import GLOBAL_TOPICS, OUTDOOR_AREA_NAME, OUTDOOR_TOPICS, ROOM_TOPICS
from app.models.project import Project
from app.services.evaluation import build_room_matrix, topic_metrics
from app.services.pricing import estimate_project_costs

HEADER_FILL = PatternFill("solid", fgColor="1D4ED8")
SECTION_FILL = PatternFill("solid", fgColor="E2E8F0")
ALT_FILL = PatternFill("solid", fgColor="F8FAFC")


def _write_topic_sheet(ws, title: str, topics, topic_values) -> None:
    ws.title = title[:31]
    ws.append(["Sektion", "Thema", "Auswahl(en)", "Notizen", "Verantwortlich"])
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = Font(color="FFFFFF", bold=True)
    row = 2
    current_section = None
    for topic in topics:
        if topic.section != current_section:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            cell = ws.cell(row=row, column=1, value=topic.section)
            cell.fill = SECTION_FILL
            cell.font = Font(bold=True)
            row += 1
            current_section = topic.section
        state = topic_values[topic.key]
        ws.append([topic.section, topic.title, ", ".join(state.selections) or "—", state.notes, state.assignee])
        if row % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = ALT_FILL
        row += 1
    for col, width in zip("ABCDE", [22, 28, 45, 48, 20]):
        ws.column_dimensions[col].width = width
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for c in r:
            c.alignment = Alignment(vertical="top", wrap_text=True)


def export_project_to_excel(project: Project, target_file: Path) -> None:
    wb = Workbook()
    ws_global = wb.active
    _write_topic_sheet(ws_global, "Global_Planung", GLOBAL_TOPICS, project.global_topics)

    ws_outdoor = wb.create_sheet(title=OUTDOOR_AREA_NAME[:31])
    _write_topic_sheet(ws_outdoor, OUTDOOR_AREA_NAME, OUTDOOR_TOPICS, project.outdoor_topics)

    for room_name, room in project.rooms.items():
        ws = wb.create_sheet(title=room_name[:31])
        _write_topic_sheet(ws, room_name, ROOM_TOPICS, room.topics)

    ws_eval = wb.create_sheet("Auswertung_Raumvergleich")
    evaluation_columns = [OUTDOOR_AREA_NAME, *project.rooms.keys()]
    ws_eval.append(["Topic", *evaluation_columns, "Bereiche mit Auswahl", "Diversity", "Dominanz"])
    for c in ws_eval[1]:
        c.fill = HEADER_FILL
        c.font = Font(color="FFFFFF", bold=True)
    matrix = build_room_matrix(project)
    metrics = topic_metrics(project)
    row = 2
    for topic, per_room in matrix.items():
        values = [", ".join(per_room[column]) or "—" for column in evaluation_columns]
        m = metrics[topic]
        ws_eval.append([
            topic,
            *values,
            f"{m['rooms_with_selection']}/{m['room_count']}",
            m["diversity"],
            round(m["dominant_ratio"], 2),
        ])
        if row % 2 == 0:
            for col in range(1, ws_eval.max_column + 1):
                ws_eval.cell(row=row, column=col).fill = ALT_FILL
        row += 1
    for i in range(1, ws_eval.max_column + 1):
        ws_eval.column_dimensions[chr(64 + i)].width = 22


    ws_cost = wb.create_sheet("Kostenübersicht")
    ws_cost.append(["Kategorie", "Beschreibung", "Menge", "Min (€)", "Typisch (€)", "Max (€)"])
    for c in ws_cost[1]:
        c.fill = HEADER_FILL
        c.font = Font(color="FFFFFF", bold=True)

    pricing = estimate_project_costs(project)
    row = 2
    for item in pricing["line_items"]:
        ws_cost.append([
            item["category"],
            item["description"],
            item["quantity"],
            item["cost"]["min"],
            item["cost"]["typical"],
            item["cost"]["max"],
        ])
        if row % 2 == 0:
            for col in range(1, 7):
                ws_cost.cell(row=row, column=col).fill = ALT_FILL
        row += 1

    totals = pricing["totals"]
    ws_cost.append(["", "", "", "", "", ""])
    ws_cost.append(["GESAMT", "", "", totals["min"], totals["typical"], totals["max"]])
    for col, width in zip("ABCDEF", [24, 50, 10, 14, 14, 14]):
        ws_cost.column_dimensions[col].width = width

    ws_cost.append(["", "", "", "", "", ""])
    ws_cost.append(["Annahmen", "", "", "", "", ""])
    for note in pricing["assumptions"]:
        ws_cost.append([f"• {note}", "", "", "", "", ""])

    target_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target_file)
